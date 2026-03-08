#!/usr/bin/env python3
"""
Build a comprehensive Auto Components sector report from all pipeline outputs.
Produces:
  - Comprehensive .md (single markdown with narrative, hypothesis, literature, universe, shortlist, backtest, stock narratives)
  - Interactive .html (tabs, sortable tables)
  - .xlsx (multiple sheets for easy consumption in Excel)
Run from project root or working-sector. Requires: pandas; for .xlsx: openpyxl (pip install openpyxl).
"""
import html as html_module
import json
import sys
from pathlib import Path
from datetime import date

import pandas as pd

WORKING_SECTOR = Path(__file__).resolve().parent
PROJECT_ROOT = WORKING_SECTOR.parent
OUTPUT_DIR = WORKING_SECTOR / "output"

# Inputs
SECTOR_NOTE_MD = OUTPUT_DIR / "auto_components_sector_note.md"
SECTOR_NARRATIVE_MD = WORKING_SECTOR / "sector_narrative_auto_components.md"
HYPOTHESIS_MD = WORKING_SECTOR / "hypothesis_auto_components.md"
LITERATURE_MD = WORKING_SECTOR / "literature_notes_auto_components.md"
PHASE3_FULL_CSV = OUTPUT_DIR / "phase3_full_with_composite.csv"
PHASE3_SHORTLIST_CSV = OUTPUT_DIR / "phase3_shortlist.csv"
PHASE4_BACKTEST_CSV = OUTPUT_DIR / "phase4_backtest_results.csv"
STOCK_NARRATIVES_JSON = OUTPUT_DIR / "stock_narratives.json"
STOCK_NARRATIVES_MD = OUTPUT_DIR / "stock_narratives.md"
UNIVERSE_CSV = WORKING_SECTOR / "auto_components_universe.csv"

# Outputs
REPORT_MD = OUTPUT_DIR / "auto_components_comprehensive_report.md"
REPORT_HTML = OUTPUT_DIR / "auto_components_comprehensive_report.html"
REPORT_XLSX = OUTPUT_DIR / "auto_components_comprehensive_report.xlsx"


def _read_text(path: Path, default: str = "") -> str:
    if path.exists():
        return path.read_text(encoding="utf-8").strip()
    return default


def _load_json(path: Path) -> list | dict:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return []


def build_md(
    sector_note: str,
    sector_narrative: str,
    hypothesis: str,
    literature: str,
    full_df: pd.DataFrame,
    shortlist_df: pd.DataFrame,
    backtest_df: pd.DataFrame,
    narratives: list[dict],
    as_of: str,
) -> str:
    today = date.today().isoformat()
    lines = [
        "# Auto Components (India) – Comprehensive Sector Report",
        "",
        f"**Report date:** {today}  |  **Data as of:** {as_of}",
        "",
        "---",
        "",
        "## 1. Sector definition and market size",
        "",
    ]
    # Use sector note narrative if available, else sector_narrative file
    narrative_block = sector_narrative or ""
    if "# " in narrative_block:
        narrative_block = "\n".join(narrative_block.split("\n")[1:]).lstrip()  # drop first heading
    lines.append(narrative_block or "*(Narrative not found.)*")
    lines += ["", "---", "", "## 2. Research question and hypothesis", ""]
    lines.append(hypothesis or "*(Hypothesis memo not found.)*")
    lines += ["", "---", "", "## 3. Literature and sources", ""]
    lines.append(literature or "*(Literature notes not found.)*")
    lines += ["", "---", "", "## 4. Universe and metrics", ""]
    lines.append(f"Component-only universe (ex-OEM), aligned with ACMA. **{len(full_df)}** stocks.")
    lines += ["", ""]

    if not full_df.empty:
        cols = ["SYMBOL", "SUBSECTOR", "CURRENT_PRICE", "RET_1M", "RET_3M", "RET_6M",
                "RS_VS_NIFTY_500_6M", "RSI", "FUND_SCORE", "TECHNICAL_SCORE", "COMPOSITE_SCORE", "PASS_SCREEN"]
        cols = [c for c in cols if c in full_df.columns]
        df = full_df[cols].copy()
        if "RET_1M" in df.columns:
            df["RET_1M"] = (df["RET_1M"] * 100).round(1).astype(str) + "%"
        if "RET_3M" in df.columns:
            df["RET_3M"] = (df["RET_3M"] * 100).round(1).astype(str) + "%"
        if "RET_6M" in df.columns:
            df["RET_6M"] = (df["RET_6M"] * 100).round(1).astype(str) + "%"
        if "RS_VS_NIFTY_500_6M" in df.columns:
            df["RS_VS_NIFTY_500_6M"] = (df["RS_VS_NIFTY_500_6M"] * 100).round(1).astype(str) + "%"
        header = "| " + " | ".join(cols) + " |"
        sep = "|" + "|".join(["---"] * len(cols)) + "|"
        lines.append(header)
        lines.append(sep)
        for _, row in df.iterrows():
            cells = [str(row.get(c, "")) for c in cols]
            lines.append("| " + " | ".join(cells) + " |")
        lines += ["", ""]

    lines += ["## 5. Shortlist (top by composite score)", ""]
    if not shortlist_df.empty:
        cols = ["SYMBOL", "SUBSECTOR", "CURRENT_PRICE", "RET_6M", "RS_VS_NIFTY_500_6M",
                "FUND_SCORE", "TECHNICAL_SCORE", "COMPOSITE_SCORE"]
        cols = [c for c in cols if c in shortlist_df.columns]
        df = shortlist_df[cols].copy()
        if "RET_6M" in df.columns:
            df["RET_6M"] = (df["RET_6M"] * 100).round(1).astype(str) + "%"
        if "RS_VS_NIFTY_500_6M" in df.columns:
            df["RS_VS_NIFTY_500_6M"] = (df["RS_VS_NIFTY_500_6M"] * 100).round(1).astype(str) + "%"
        header = "| " + " | ".join(cols) + " |"
        sep = "|" + "|".join(["---"] * len(cols)) + "|"
        lines.append(header)
        lines.append(sep)
        for _, row in df.iterrows():
            cells = [str(row.get(c, "")) for c in cols]
            lines.append("| " + " | ".join(cells) + " |")
    lines += ["", "---", "", "## 6. Backtest (momentum screen: RS_6M > 0)", ""]
    if not backtest_df.empty and "EXCESS_RET" in backtest_df.columns:
        excess = backtest_df["EXCESS_RET"].dropna()
        mean_excess = excess.mean() * 100 if len(excess) > 0 else float("nan")
        hit = (excess > 0).sum() / len(excess) * 100 if len(excess) > 0 else 0
        lines.append(f"- Mean excess return (portfolio vs Nifty 500, 1Y forward): **{mean_excess:.2f}%**")
        lines.append(f"- Hit rate (excess > 0): **{hit:.0f}%**")
        lines.append("")
        lines.append("*Backtest uses only price-based criteria; fundamental screen not applied historically.*")
    lines += ["", "---", "", "## 7. Stock narratives (with fundamental details)", ""]
    lines.append("Per-stock narratives, fundamental scores (0–100), and when available: P&L (Sales, Net Profit, EPS, YoY), quarterly results, balance sheet (Equity, Debt, Cash, D/E), and ratios (ROCE, ROE, EPS, PE, PB, OPM, NPM, etc.).")
    lines += ["", ""]
    for n in narratives:
        sym, name = n.get("symbol", ""), n.get("name", "")
        lines.append(f"### {sym} – {name}")
        lines.append("")
        fund_parts = []
        for k in ["fund_earnings_quality", "fund_sales_growth", "fund_financial_strength", "fund_institutional_backing"]:
            if k in n and n[k] is not None:
                label = k.replace("fund_", "").replace("_", " ").title()
                fund_parts.append(f"**{label}:** {n[k]:.1f}")
        if fund_parts:
            lines.append("*Fundamental (0–100):* " + " | ".join(fund_parts))
            lines.append("")
        for key, label in [
            ("pnl_summary", "P&L"),
            ("quarterly_summary", "Quarterly"),
            ("balance_sheet_summary", "Balance sheet"),
            ("ratios_summary", "Ratios"),
        ]:
            if n.get(key) and str(n[key]).strip():
                lines.append(f"*{label}:* " + str(n[key]).strip())
                lines.append("")
        lines.append(n.get("narrative", ""))
        lines += ["", ""]
    lines += ["---", "", "## 8. Sources and data", ""]
    lines.append("- **Definition and market size:** ACMA FY25; listed universe and mcap from [Sharescart Auto Ancillary](https://www.sharescart.com/industry/auto-ancillary/) (113 companies, ₹7.53 L Cr); sector_view_auto_components.md, literature_notes_auto_components.md.")
    lines.append("- **Strategic outlook:** auto-components.md (policy, EV, ADAS, clusters, challenges).")
    lines.append("- **Price data:** NSE (nse_sec_full_data.csv, nse_index_data.csv).")
    lines.append("- **Fundamental scores:** Screener/organized pipeline (fundamental_scores_database.csv).")
    lines.append("- **Narratives:** Generated using Ollama Granite 4 from pipeline, fundamental scores, and optional P&L/quarterly/balance-sheet/ratios (ROCE, ROE, EPS, PE, PB, etc.) from fetch_screener_fundamental_details.R.")
    return "\n".join(lines)


def _df_to_html_table(
    df: pd.DataFrame,
    table_id: str,
    class_name: str = "sortable",
    raw_df: pd.DataFrame | None = None,
    highlight_top_cols: list[str] | None = None,
    top_frac: float = 0.25,
) -> str:
    """Build sortable table HTML. If raw_df and highlight_top_cols given, add cell-best to top fraction of numeric cols."""
    if df is None or df.empty:
        return "<p>No data.</p>"
    cols = list(df.columns)
    # Precompute which (row_idx, col_name) are in top fraction (higher is better)
    best_cells: set[tuple[int, str]] = set()
    if raw_df is not None and highlight_top_cols and not raw_df.empty:
        for c in highlight_top_cols:
            if c not in raw_df.columns:
                continue
            s = pd.to_numeric(raw_df[c], errors="coerce").dropna()
            if s.empty or len(s) < 2:
                continue
            thresh = s.quantile(1 - top_frac)
            for pos, idx in enumerate(raw_df.index):
                v = raw_df.loc[idx, c]
                try:
                    if pd.notna(v) and float(v) >= thresh:
                        best_cells.add((pos, c))
                except (TypeError, ValueError):
                    pass
    html = [f'<table id="{table_id}" class="{class_name}"><thead><tr>']
    for c in cols:
        html.append(f'<th data-col="{html_module.escape(c)}">{c}</th>')
    html.append("</tr></thead><tbody>")
    for row_idx, (_, row) in enumerate(df.iterrows()):
        html.append("<tr>")
        for c in cols:
            v = row.get(c, "")
            if pd.isna(v):
                v = ""
            elif isinstance(v, float):
                if "RET" in c or "RS_" in c or "EXCESS" in c:
                    v = f"{v*100:.1f}%" if abs(v) < 10 else f"{v:.2f}"
                else:
                    v = f"{v:.2f}" if abs(v) < 1e6 else f"{v:.1f}"
            cls = " cell-best" if (row_idx, c) in best_cells else ""
            html.append(f"<td class=\"{cls.strip()}\">{v}</td>")
        html.append("</tr>")
    html.append("</tbody></table>")
    return "\n".join(html)


def _markdown_para_to_html(para: str) -> str:
    """Convert **bold** (and *italic*) in a paragraph to HTML, with escaping for safety."""
    if not para or not para.strip():
        return ""
    parts = para.split("**")
    out = []
    for i, p in enumerate(parts):
        if i % 2 == 1:
            out.append("<strong>" + html_module.escape(p) + "</strong>")
        else:
            # Support *italic* within non-bold segments
            sub_parts = p.split("*")
            for j, sp in enumerate(sub_parts):
                if j % 2 == 1:
                    out.append("<em>" + html_module.escape(sp) + "</em>")
                else:
                    out.append(html_module.escape(sp))
    return "".join(out)


def _narratives_to_html(narratives: list[dict], html_escape_fn) -> str:
    """Render narratives as cards: heading, fundamental table, narrative paragraphs (markdown rendered)."""
    if not narratives:
        return "<p>No narratives.</p>"
    fund_labels = {
        "fund_earnings_quality": "Earnings quality",
        "fund_sales_growth": "Sales growth",
        "fund_financial_strength": "Financial strength",
        "fund_institutional_backing": "Institutional backing",
    }
    out = []
    for n in narratives:
        sym = n.get("symbol", "")
        name = n.get("name", "")
        out.append(f'<div class="narrative-card">')
        out.append(f'<h3 class="narrative-title">{html_escape_fn(sym)} – {html_escape_fn(name)}</h3>')
        if n.get("recommendation"):
            rec = html_escape_fn(str(n["recommendation"]))
            rationale = html_escape_fn(str(n.get("recommendation_rationale", "")).strip())
            out.append(f'<p class="narrative-recommendation"><strong>Recommendation:</strong> {rec}')
            if rationale:
                out.append(f' <span class="narrative-rec-rationale">— {rationale}</span>')
            out.append('</p>')
        fund_keys = sorted([k for k in n if k.startswith("fund_") and n.get(k) is not None])
        if fund_keys:
            labels = [fund_labels.get(k, k.replace("fund_", "").replace("_", " ").title()) for k in fund_keys]
            values = [f"{n[k]:.1f}" for k in fund_keys]
            out.append('<p class="narrative-fund-label"><em>Fundamental scores (0–100)</em></p>')
            out.append('<table class="fund-table-mini"><thead><tr>')
            for L in labels:
                out.append(f"<th>{html_escape_fn(L)}</th>")
            out.append("</tr></thead><tbody><tr>")
            for v in values:
                out.append(f"<td>{v}</td>")
            out.append("</tr></tbody></table>")
        for key, label in [
            ("pnl_summary", "P&L (key items, YoY)"),
            ("quarterly_summary", "Quarterly results"),
            ("balance_sheet_summary", "Balance sheet"),
            ("ratios_summary", "Ratios (ROCE, ROE, EPS, PE, PB, etc.)"),
        ]:
            if n.get(key) and str(n[key]).strip():
                out.append(f'<p class="narrative-fund-label"><em>{html_escape_fn(label)}</em></p>')
                out.append(f'<p class="narrative-para narrative-extra">{html_escape_fn(str(n[key]).strip())}</p>')
        narrative = (n.get("narrative") or "").strip()
        if narrative:
            for para in narrative.split("\n\n"):
                para = para.strip()
                if para:
                    out.append(f"<p class=\"narrative-para\">{_markdown_para_to_html(para)}</p>")
        out.append("</div>")
    return "\n".join(out)


def _guide_tab_html() -> str:
    """HTML for the 'How to read' tab: scores, metrics, and report structure."""
    return """
<div class="guide-content">
  <section class="guide-section">
    <h3>How to use this report</h3>
    <p>This report has six tabs: <strong>Overview</strong> (sector definition, hypothesis, literature), <strong>How to read</strong> (this guide), <strong>Universe</strong> (all stocks with metrics), <strong>Shortlist</strong> (top names by composite score), <strong>Backtest</strong> (historical momentum screen results), and <strong>Stock narratives</strong> (per-stock commentary with fundamental breakdown). Use the tabs to switch views.</p>
    <p><strong>Tables:</strong> In Universe, Shortlist, and Backtest, click any column header to sort by that column; click again to toggle ascending (▲) / descending (▼). <strong>Green-highlighted cells</strong> mark the top 25% of values in that column (e.g. highest composite score, best 6M return, highest excess return in backtest).</p>
  </section>

  <section class="guide-section">
    <h3>Price and return metrics</h3>
    <table class="guide-table">
      <thead><tr><th>Metric</th><th>Description</th></tr></thead>
      <tbody>
        <tr><td><strong>CURRENT_PRICE</strong></td><td>Latest closing price (Rs).</td></tr>
        <tr><td><strong>RET_1M / RET_3M / RET_6M</strong></td><td>Total return over 1, 3, and 6 months (e.g. 10% = 0.10).</td></tr>
        <tr><td><strong>RS_VS_NIFTY_500_6M</strong></td><td>Stock return minus Nifty 500 return over 6 months. Positive = outperformed the index.</td></tr>
        <tr><td><strong>RS_VS_NIFTY_AUTO_6M</strong></td><td>Same vs Nifty Auto index.</td></tr>
      </tbody>
    </table>
  </section>

  <section class="guide-section">
    <h3>Technical metrics</h3>
    <table class="guide-table">
      <thead><tr><th>Metric</th><th>Description</th></tr></thead>
      <tbody>
        <tr><td><strong>RSI</strong></td><td>Relative Strength Index (0–100). &lt;30 often oversold, &gt;70 overbought.</td></tr>
        <tr><td><strong>TECHNICAL_SCORE</strong></td><td>Composite technical score (e.g. 0–100) from price and momentum; higher = stronger short-term technical picture.</td></tr>
      </tbody>
    </table>
  </section>

  <section class="guide-section">
    <h3>Fundamental scores</h3>
    <p>All fundamental scores are on a 0–100 scale (higher = better). They come from the Screener/organized pipeline.</p>
    <table class="guide-table">
      <thead><tr><th>Score</th><th>Description</th></tr></thead>
      <tbody>
        <tr><td><strong>FUND_SCORE</strong></td><td>Overall fundamental score (earnings quality, sales growth, financial strength, institutional backing combined).</td></tr>
        <tr><td><strong>Earnings quality</strong></td><td>Quality and sustainability of earnings (margins, consistency, cash flow).</td></tr>
        <tr><td><strong>Sales growth</strong></td><td>Revenue growth trajectory.</td></tr>
        <tr><td><strong>Financial strength</strong></td><td>Balance sheet health, leverage, liquidity.</td></tr>
        <tr><td><strong>Institutional backing</strong></td><td>Degree of institutional (DII/FII) ownership and trend.</td></tr>
      </tbody>
    </table>
  </section>

  <section class="guide-section">
    <h3>Composite and screen</h3>
    <table class="guide-table">
      <thead><tr><th>Term</th><th>Description</th></tr></thead>
      <tbody>
        <tr><td><strong>COMPOSITE_SCORE</strong></td><td>Weighted combination: 0.4 × fundamental + 0.4 × technical + 0.2 × relative strength rank (0–100). Used to rank and shortlist stocks.</td></tr>
        <tr><td><strong>PASS_FUND</strong></td><td>True if fundamental score &gt;= 70 (quality screen).</td></tr>
        <tr><td><strong>PASS_RS</strong></td><td>True if 6M relative strength vs Nifty 500 &gt; 0 (momentum screen).</td></tr>
        <tr><td><strong>PASS_SCREEN</strong></td><td>True if both PASS_FUND and PASS_RS.</td></tr>
      </tbody>
    </table>
  </section>

  <section class="guide-section">
    <h3>Backtest</h3>
    <p><strong>REBAL_DATE</strong>: Date when the portfolio was formed. <strong>N_STOCKS</strong>: Number of stocks passing the screen. <strong>PORTFOLIO_RET_1Y</strong>: Equal-weight return of those stocks over the next year. <strong>BENCH_RET_1Y</strong>: Nifty 500 return over the same period. <strong>EXCESS_RET</strong>: Portfolio return minus benchmark. The backtest uses a momentum-only screen (RS &gt; 0); fundamental filters are not applied historically due to data limits.</p>
  </section>

  <section class="guide-section">
    <h3>Stock narratives</h3>
    <p>Each stock has a short narrative that synthesizes <strong>key metrics</strong> (price, returns, RSI, technical and fundamental scores) and <strong>key financial ratios/fundamentals</strong>: P&amp;L (Sales, Net Profit, EPS, YoY), <strong>Quarterly</strong> (last 4Q sales and profit), <strong>Balance sheet</strong> (Equity, Debt, Cash, Debt/Equity), and <strong>Ratios</strong> (ROCE, ROE, EPS, PE, PB, OPM, NPM, Interest Coverage, Div Yield when available). The <strong>fundamental score (0–100)</strong> is the main financial-quality metric. Run <code>fetch_screener_fundamental_details.R</code> to populate these from Screener.in.</p>
    <p><strong>Recommendation:</strong> The LLM (Ollama) provides a view-based recommendation (Buy / Accumulate / Hold / Reduce / Avoid) with a one-line rationale derived from the metrics. This is for research context only, not investment advice.</p>
  </section>
</div>
"""


def build_html(
    sector_narrative: str,
    hypothesis: str,
    literature: str,
    full_df: pd.DataFrame,
    shortlist_df: pd.DataFrame,
    backtest_df: pd.DataFrame,
    narratives: list[dict],
    as_of: str,
) -> str:
    today = date.today().isoformat()
    # Prepare display dataframes (percent formatting)
    full_display = full_df.copy() if full_df is not None and not full_df.empty else pd.DataFrame()
    for c in ["RET_1M", "RET_3M", "RET_6M", "RS_VS_NIFTY_AUTO_6M", "RS_VS_NIFTY_500_6M"]:
        if c in full_display.columns:
            full_display[c] = (full_display[c] * 100).round(1).astype(str) + "%"
    shortlist_display = shortlist_df.copy() if shortlist_df is not None and not shortlist_df.empty else pd.DataFrame()
    for c in ["RET_6M", "RS_VS_NIFTY_500_6M"]:
        if c in shortlist_display.columns:
            shortlist_display[c] = (shortlist_display[c] * 100).round(1).astype(str) + "%"
    backtest_display = backtest_df.copy() if backtest_df is not None and not backtest_df.empty else pd.DataFrame()
    for c in ["PORTFOLIO_RET_1Y", "BENCH_RET_1Y", "EXCESS_RET"]:
        if c in backtest_display.columns:
            backtest_display[c] = (backtest_display[c] * 100).round(2).astype(str) + "%"

    narratives_df = pd.DataFrame(narratives) if narratives else pd.DataFrame()
    if not narratives_df.empty:
        base_cols = [c for c in ["symbol", "name", "narrative"] if c in narratives_df.columns]
        fund_cols = [c for c in narratives_df.columns if c.startswith("fund_")]
        narratives_df = narratives_df[base_cols + fund_cols] if (base_cols or fund_cols) else narratives_df

    def html_escape(s: str) -> str:
        if not s:
            return ""
        return (
            s.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
        )

    # Overview: raw markdown for client-side rendering (marked.js); truncate for display
    overview_narrative = (sector_narrative[:3000] + "…") if sector_narrative and len(sector_narrative) > 3000 else (sector_narrative or "")
    overview_hypothesis = (hypothesis[:4000] + "…") if hypothesis and len(hypothesis) > 4000 else (hypothesis or "")
    overview_literature = (literature[:5000] + "…") if literature and len(literature) > 5000 else (literature or "")
    overview_json = json.dumps(
        {"narrative": overview_narrative, "hypothesis": overview_hypothesis, "literature": overview_literature},
        ensure_ascii=False,
    ).replace("</script>", "<\\/script>")  # avoid breaking script tag in HTML

    narratives_html = _narratives_to_html(narratives, html_escape)
    guide_html = _guide_tab_html()

    tabs = [
        ("overview", "Overview", (
            "<div class='overview-cards'>"
            "<div class='card-section'><h3 class='card-title'>Sector definition &amp; market size</h3>"
            "<div id='overview-narrative' class='md-rendered'></div></div>"
            "<div class='card-section'><h3 class='card-title'>Research question &amp; hypothesis</h3>"
            "<div id='overview-hypothesis' class='md-rendered'></div></div>"
            "<div class='card-section'><h3 class='card-title'>Literature &amp; sources</h3>"
            "<div id='overview-literature' class='md-rendered'></div></div>"
            "</div>"
            f"<script type='application/json' id='overview-md'>{overview_json}</script>"
        )),
        ("guide", "How to read", guide_html),
        ("universe", "Universe", '<div class="table-scroll">' + _df_to_html_table(
            full_display, "table-universe",
            raw_df=full_df, highlight_top_cols=["COMPOSITE_SCORE", "FUND_SCORE", "RET_6M", "RS_VS_NIFTY_500_6M", "TECHNICAL_SCORE"],
        ) + '</div>'),
        ("shortlist", "Shortlist", '<div class="table-scroll">' + _df_to_html_table(
            shortlist_display, "table-shortlist",
            raw_df=shortlist_df, highlight_top_cols=["COMPOSITE_SCORE", "FUND_SCORE", "RET_6M", "RS_VS_NIFTY_500_6M", "TECHNICAL_SCORE"],
        ) + '</div>'),
        ("backtest", "Backtest", '<div class="table-scroll">' + _df_to_html_table(
            backtest_display, "table-backtest",
            raw_df=backtest_df, highlight_top_cols=["EXCESS_RET", "PORTFOLIO_RET_1Y"],
        ) + '</div>'),
        ("narratives", "Stock narratives", narratives_html),
    ]

    tab_buttons = "".join(
        f'<button class="tab-btn" data-tab="{tid}">{label}</button>' for tid, label, _ in tabs
    )
    tab_panels = "".join(
        f'<div id="panel-{tid}" class="tab-panel" role="tabpanel">{content}</div>'
        for tid, _, content in tabs
    )

    return """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Auto Components (India) – Comprehensive Report</title>
<link href="https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&display=swap" rel="stylesheet">
<style>
:root {
  --md-primary: #00897B;
  --md-primary-dark: #00695C;
  --md-primary-light: #4DB6AC;
  --md-primary-bg: #B2DFDB;
  --md-surface: #FFFFFF;
  --md-bg: #E0F2F1;
  --md-text: #004D40;
  --md-text-secondary: #00796B;
  --md-divider: rgba(0, 137, 123, 0.2);
  --md-elevation-1: 0 1px 3px rgba(0, 105, 92, 0.12), 0 1px 2px rgba(0, 105, 92, 0.24);
  --md-elevation-2: 0 3px 6px rgba(0, 105, 92, 0.15), 0 2px 4px rgba(0, 105, 92, 0.12);
  --md-elevation-3: 0 10px 20px rgba(0, 105, 92, 0.15), 0 3px 6px rgba(0, 105, 92, 0.1);
  --md-radius: 8px;
  --md-radius-sm: 4px;
}
* { box-sizing: border-box; }
body {
  font-family: 'Roboto', system-ui, sans-serif;
  margin: 0;
  padding: 0;
  background: var(--md-bg);
  color: var(--md-text);
  font-size: 14px;
  line-height: 1.5;
  -webkit-font-smoothing: antialiased;
}
.app-bar {
  background: linear-gradient(135deg, var(--md-primary-dark) 0%, var(--md-primary) 100%);
  color: #fff;
  padding: 16px 24px;
  box-shadow: var(--md-elevation-2);
}
.app-bar h1 {
  margin: 0;
  font-size: 1.5rem;
  font-weight: 500;
  letter-spacing: 0.02em;
}
.app-bar .meta {
  margin: 4px 0 0 0;
  font-size: 0.875rem;
  opacity: 0.9;
  font-weight: 300;
}
.main-content { padding: 24px; max-width: 1400px; margin: 0 auto; }
.tabs {
  display: flex;
  gap: 0;
  margin-bottom: 0;
  background: var(--md-surface);
  border-radius: var(--md-radius) var(--md-radius) 0 0;
  box-shadow: var(--md-elevation-1);
  overflow: hidden;
  flex-wrap: wrap;
}
.tab-btn {
  padding: 14px 20px;
  cursor: pointer;
  border: none;
  background: transparent;
  font-family: inherit;
  font-size: 0.875rem;
  font-weight: 500;
  color: var(--md-text-secondary);
  text-transform: uppercase;
  letter-spacing: 0.05em;
  position: relative;
  transition: color 0.2s, background 0.2s;
}
.tab-btn:hover { background: var(--md-bg); color: var(--md-primary-dark); }
.tab-btn.active {
  color: var(--md-primary);
  background: rgba(0, 137, 123, 0.08);
}
.tab-btn.active::after {
  content: '';
  position: absolute;
  bottom: 0;
  left: 0;
  right: 0;
  height: 3px;
  background: var(--md-primary);
}
.tab-panel {
  display: none;
  background: var(--md-surface);
  padding: 24px;
  border-radius: 0 0 var(--md-radius) var(--md-radius);
  box-shadow: var(--md-elevation-1);
  overflow-x: auto;
  border: 1px solid var(--md-divider);
  border-top: none;
  min-height: 320px;
}
.tab-panel.active { display: block; }
.card-section {
  background: var(--md-surface);
  border-radius: var(--md-radius);
  padding: 20px;
  margin-bottom: 20px;
  box-shadow: var(--md-elevation-1);
  border: 1px solid var(--md-divider);
}
.overview-cards { display: flex; flex-direction: column; gap: 20px; }
.card-section pre {
  font-family: 'Roboto', sans-serif;
  white-space: pre-wrap;
  word-wrap: break-word;
  margin: 0;
  padding: 16px;
  background: var(--md-bg);
  border-radius: var(--md-radius-sm);
  border-left: 4px solid var(--md-primary-light);
  font-size: 0.8125rem;
  line-height: 1.6;
  color: var(--md-text);
}
.card-title {
  margin: 0 0 12px 0;
  font-size: 1rem;
  font-weight: 500;
  color: var(--md-primary-dark);
  padding-bottom: 8px;
  border-bottom: 1px solid var(--md-divider);
}
.md-rendered { font-size: 0.9rem; line-height: 1.6; }
.md-rendered h1 { font-size: 1.1rem; margin: 0 0 0.5rem 0; color: var(--md-primary-dark); font-weight: 500; }
.md-rendered h2 { font-size: 1rem; margin: 1rem 0 0.4rem 0; color: var(--md-primary-dark); font-weight: 500; }
.md-rendered h3 { font-size: 0.95rem; margin: 0.75rem 0 0.35rem 0; color: var(--md-primary-dark); font-weight: 500; }
.md-rendered p { margin: 0 0 0.6rem 0; }
.md-rendered ul, .md-rendered ol { margin: 0.4rem 0 0.75rem 0; padding-left: 1.5rem; }
.md-rendered li { margin: 0.2rem 0; }
.md-rendered hr { border: none; border-top: 1px solid var(--md-divider); margin: 1rem 0; }
.md-rendered strong { font-weight: 500; color: var(--md-text); }
.md-rendered table { width: 100%; margin: 0.5rem 0; font-size: 0.8rem; border-collapse: collapse; }
.md-rendered th, .md-rendered td { border: 1px solid var(--md-divider); padding: 6px 8px; text-align: left; }
.md-rendered th { background: var(--md-primary-light); color: var(--md-text); font-weight: 500; }
.md-rendered a { color: var(--md-primary); text-decoration: none; }
.md-rendered a:hover { text-decoration: underline; }
.md-rendered.md-fallback { white-space: pre-wrap; font-family: inherit; }
.prose { padding: 0 4px; }
.prose h3 {
  margin: 24px 0 12px 0;
  font-size: 1.125rem;
  font-weight: 500;
  color: var(--md-primary-dark);
  padding-bottom: 8px;
  border-bottom: 1px solid var(--md-divider);
}
.prose h3:first-child { margin-top: 0; }
.prose pre {
  font-family: 'Roboto', sans-serif;
  white-space: pre-wrap;
  word-wrap: break-word;
  margin: 0 0 16px 0;
  padding: 16px;
  background: var(--md-bg);
  border-radius: var(--md-radius-sm);
  border-left: 4px solid var(--md-primary-light);
  font-size: 0.8125rem;
  line-height: 1.6;
  color: var(--md-text);
}
table {
  border-collapse: collapse;
  width: 100%;
  font-size: 0.8125rem;
  border-radius: var(--md-radius);
  overflow: hidden;
  box-shadow: var(--md-elevation-1);
}
th, td {
  padding: 12px 16px;
  text-align: left;
  border-bottom: 1px solid var(--md-divider);
}
th {
  background: var(--md-primary);
  color: #fff;
  cursor: pointer;
  user-select: none;
  white-space: nowrap;
  font-weight: 500;
  text-transform: uppercase;
  font-size: 0.75rem;
  letter-spacing: 0.05em;
  transition: background 0.2s;
}
th:hover { background: var(--md-primary-dark); }
.sortable th { cursor: pointer; user-select: none; position: relative; padding-right: 24px; }
.sortable th::after { content: ' ⇅'; opacity: 0.4; font-size: 0.85em; }
.sortable th.sort-asc::after { content: ' ▲'; opacity: 1; }
.sortable th.sort-desc::after { content: ' ▼'; opacity: 1; }
tbody tr {
  transition: background 0.15s;
}
tbody tr:nth-child(even) { background: rgba(0, 137, 123, 0.04); }
tbody tr:hover { background: rgba(0, 137, 123, 0.08); }
.sortable { border-collapse: collapse; width: 100%; border-radius: var(--md-radius); overflow: hidden; box-shadow: var(--md-elevation-1); }
.sortable td, .sortable th { border-bottom: 1px solid var(--md-divider); }
.sortable tbody tr:last-child td { border-bottom: none; }
.cell-best { background: rgba(0, 150, 136, 0.18) !important; font-weight: 600; color: var(--md-primary-dark); }
.table-scroll { overflow-x: auto; margin: 0 -4px; border-radius: var(--md-radius); }
#table-narratives td { max-width: 420px; word-wrap: break-word; }
.narrative-card {
  background: var(--md-surface);
  border: 1px solid var(--md-divider);
  border-radius: var(--md-radius);
  padding: 20px;
  margin-bottom: 24px;
  box-shadow: var(--md-elevation-1);
}
.narrative-title {
  margin: 0 0 8px 0;
  font-size: 1.1rem;
  font-weight: 500;
  color: var(--md-primary-dark);
  padding-bottom: 8px;
  border-bottom: 1px solid var(--md-divider);
}
.narrative-recommendation {
  margin: 0 0 12px 0;
  padding: 8px 12px;
  background: var(--md-bg);
  border-left: 4px solid var(--md-primary);
  font-size: 0.95rem;
}
.narrative-rec-rationale {
  color: var(--md-text-secondary);
  font-style: italic;
}
.narrative-fund-label { margin: 12px 0 6px 0; font-size: 0.875rem; color: var(--md-text-secondary); }
.fund-table-mini {
  width: auto;
  min-width: 320px;
  margin-bottom: 12px;
  font-size: 0.8125rem;
}
.fund-table-mini th { padding: 8px 12px; }
.fund-table-mini td { padding: 8px 12px; }
.narrative-para {
  margin: 0 0 12px 0;
  line-height: 1.6;
  color: var(--md-text);
}
.narrative-para:last-of-type { margin-bottom: 0; }
.guide-content { max-width: 720px; }
.guide-section {
  margin-bottom: 24px;
}
.guide-section h3 {
  font-size: 1rem;
  font-weight: 500;
  color: var(--md-primary-dark);
  margin: 0 0 10px 0;
  padding-bottom: 6px;
  border-bottom: 1px solid var(--md-divider);
}
.guide-section p {
  margin: 0 0 12px 0;
  line-height: 1.6;
}
.guide-table {
  width: 100%;
  margin-top: 8px;
  margin-bottom: 16px;
  font-size: 0.875rem;
}
.guide-table th { background: var(--md-primary-light); color: var(--md-text); }
@media (max-width: 600px) {
  .app-bar { padding: 12px 16px; }
  .main-content { padding: 16px; }
  .tab-btn { padding: 12px 14px; font-size: 0.75rem; }
  .tab-panel { padding: 16px; }
}
</style>
</head>
<body>
<header class="app-bar">
  <h1>Auto Components (India) – Comprehensive Sector Report</h1>
  <p class="meta">Report date: """ + today + """ &nbsp;·&nbsp; Data as of: """ + as_of + """</p>
</header>
<main class="main-content">
  <div class="tabs" role="tablist">""" + tab_buttons + """</div>
""" + tab_panels + """
</main>
<script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
<script>
(function() {
  var renderOverviewMd = function() {
    var el = document.getElementById('overview-md');
    if (!el) return;
    try {
      var data = JSON.parse(el.textContent);
      var ids = ['narrative', 'hypothesis', 'literature'];
      ids.forEach(function(key) {
        var div = document.getElementById('overview-' + key);
        if (!div) return;
        var md = data[key] || '';
        if (typeof marked !== 'undefined' && marked.parse) div.innerHTML = marked.parse(md);
        else { div.textContent = md; div.classList.add('md-fallback'); }
      });
    } catch (e) {}
  };
  if (document.readyState === 'loading') document.addEventListener('DOMContentLoaded', renderOverviewMd);
  else renderOverviewMd();

  var panels = document.querySelectorAll('.tab-panel');
  var buttons = document.querySelectorAll('.tab-btn');
  function show(id) {
    panels.forEach(function(p) { p.classList.remove('active'); });
    buttons.forEach(function(b) { b.classList.remove('active'); });
    var p = document.getElementById('panel-' + id);
    var b = document.querySelector('[data-tab="' + id + '"]');
    if (p) p.classList.add('active');
    if (b) b.classList.add('active');
  }
  buttons.forEach(function(b) {
    b.addEventListener('click', function() { show(b.getAttribute('data-tab')); });
  });
  if (buttons.length) show(buttons[0].getAttribute('data-tab'));

  function parseCellNum(txt) {
    if (!txt) return NaN;
    var s = txt.replace(/%|,/g, '').trim();
    var n = parseFloat(s);
    return isNaN(n) ? NaN : n;
  }
  function sortTable(tableId) {
    var table = document.getElementById(tableId);
    if (!table) return;
    var headers = table.querySelectorAll('thead th');
    var body = table.querySelector('tbody');
    var sortState = { colIndex: -1, asc: true };
    headers.forEach(function(th, colIndex) {
      th.addEventListener('click', function() {
        var rows = Array.from(body.querySelectorAll('tr'));
        if (!rows.length) return;
        var asc = (sortState.colIndex === colIndex) ? !sortState.asc : true;
        sortState.colIndex = colIndex;
        sortState.asc = asc;
        headers.forEach(function(h, i) {
          h.classList.remove('sort-asc', 'sort-desc');
          if (i === colIndex) h.classList.add(asc ? 'sort-asc' : 'sort-desc');
        });
        var cellVal = function(row, ci) {
          var c = row.cells[ci];
          return c ? c.textContent.trim() : '';
        };
        var isNum = false;
        for (var r = 0; r < rows.length; r++) {
          var n = parseCellNum(cellVal(rows[r], colIndex));
          if (!isNaN(n)) { isNum = true; break; }
        }
        rows.sort(function(a, b) {
          var va = cellVal(a, colIndex);
          var vb = cellVal(b, colIndex);
          if (isNum) {
            var na = parseCellNum(va);
            var nb = parseCellNum(vb);
            if (isNaN(na)) na = -Infinity;
            if (isNaN(nb)) nb = -Infinity;
            return asc ? na - nb : nb - na;
          }
          return asc ? va.localeCompare(vb, undefined, { numeric: true }) : vb.localeCompare(va, undefined, { numeric: true });
        });
        rows.forEach(function(r) { body.appendChild(r); });
      });
    });
  }
  ['table-universe','table-shortlist','table-backtest'].forEach(sortTable);
})();
</script>
</body>
</html>"""


def _markdown_to_plain(s: str) -> str:
    """Convert markdown to plain text for Excel: strip bold/italic, headings, table pipes."""
    if not s or not s.strip():
        return s or ""
    import re
    t = s.strip()
    t = re.sub(r"\*\*(.+?)\*\*", r"\1", t)  # **bold** -> bold
    t = re.sub(r"\*(.+?)\*", r"\1", t)       # *italic* -> italic
    t = re.sub(r"^#+\s*", "", t, flags=re.MULTILINE)  # # heading -> heading
    t = re.sub(r"^---+$", "", t, flags=re.MULTILINE)  # ---
    t = re.sub(r"^\s*[-*]\s+", "  • ", t, flags=re.MULTILINE)  # - item -> • item
    t = re.sub(r"\n{3,}", "\n\n", t)
    t = re.sub(r"^\|\s*", "", t, flags=re.MULTILINE)
    t = re.sub(r"\s*\|$", "", t, flags=re.MULTILINE)
    t = re.sub(r"\s*\|\s*", " | ", t)
    t = re.sub(r"^\|?[\s\-:|]+\|?\s*$", "", t, flags=re.MULTILINE)  # table separator row
    t = re.sub(r"\n{3,}", "\n\n", t).strip()
    return t


def _guide_sheet_data() -> list[list[str]]:
    """Rows for the Guide sheet: metrics and how to read the report."""
    return [
        ["How to use this report", ""],
        ["Tabs / Sheets", "Summary = key dates and counts. Sector_narrative, Hypothesis, Literature = text. Universe_metrics = all stocks. Shortlist = top by composite. Backtest = momentum screen history. Stock_narratives = per-stock commentary and fundamental scores. Guide = this sheet."],
        ["", ""],
        ["Price and return metrics", ""],
        ["CURRENT_PRICE", "Latest closing price (Rs)."],
        ["RET_1M / RET_3M / RET_6M", "Total return over 1, 3, 6 months (e.g. 0.10 = 10%)."],
        ["RS_VS_NIFTY_500_6M", "Stock return minus Nifty 500 over 6M. Positive = outperformed."],
        ["", ""],
        ["Technical", ""],
        ["RSI", "0–100. <30 often oversold, >70 overbought."],
        ["TECHNICAL_SCORE", "Composite technical score; higher = stronger."],
        ["", ""],
        ["Fundamental (0–100)", ""],
        ["FUND_SCORE", "Overall (earnings quality, sales growth, financial strength, institutional backing)."],
        ["Earnings quality", "Quality and sustainability of earnings."],
        ["Sales growth", "Revenue growth trajectory."],
        ["Financial strength", "Balance sheet health, leverage."],
        ["Institutional backing", "DII/FII ownership and trend."],
        ["", ""],
        ["Composite and screen", ""],
        ["COMPOSITE_SCORE", "0.4×fundamental + 0.4×technical + 0.2×RS rank. Used to rank/shortlist."],
        ["PASS_FUND", "True if fundamental score >= 70."],
        ["PASS_RS", "True if 6M relative strength vs Nifty 500 > 0."],
        ["PASS_SCREEN", "True if both PASS_FUND and PASS_RS."],
        ["", ""],
        ["Backtest", ""],
        ["REBAL_DATE", "Date portfolio formed. PORTFOLIO_RET_1Y = equal-weight return. BENCH_RET_1Y = Nifty 500. EXCESS_RET = difference. Screen is momentum-only (RS>0)."],
    ]


def build_xlsx(
    sector_narrative: str,
    hypothesis: str,
    literature: str,
    full_df: pd.DataFrame,
    shortlist_df: pd.DataFrame,
    backtest_df: pd.DataFrame,
    narratives: list[dict],
    as_of: str,
) -> None:
    try:
        import openpyxl  # noqa: F401
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  Skipping .xlsx: install openpyxl (pip install openpyxl)")
        return
    today = date.today().isoformat()
    narrative_plain = _markdown_to_plain(sector_narrative or "")
    hypothesis_plain = _markdown_to_plain(hypothesis or "")
    literature_plain = _markdown_to_plain(literature or "")

    with pd.ExcelWriter(REPORT_XLSX, engine="openpyxl") as writer:
        summary = pd.DataFrame([
            ["Report date", today],
            ["Data as of", as_of],
            ["Sector", "Auto Components (India)"],
            ["Universe count", len(full_df) if full_df is not None and not full_df.empty else 0],
            ["Shortlist count", len(shortlist_df) if shortlist_df is not None and not shortlist_df.empty else 0],
        ])
        summary.to_excel(writer, sheet_name="Summary", index=False, header=False)
        narrative_df = pd.DataFrame([["Sector definition and market size", narrative_plain or "(none)"]])
        narrative_df.to_excel(writer, sheet_name="Sector_narrative", index=False, header=False)
        hyp_df = pd.DataFrame([["Research question and hypothesis", hypothesis_plain or "(none)"]])
        hyp_df.to_excel(writer, sheet_name="Hypothesis", index=False, header=False)
        lit_df = pd.DataFrame([["Literature and sources", literature_plain or "(none)"]])
        lit_df.to_excel(writer, sheet_name="Literature", index=False, header=False)
        if full_df is not None and not full_df.empty:
            full_df.to_excel(writer, sheet_name="Universe_metrics", index=False)
        if shortlist_df is not None and not shortlist_df.empty:
            shortlist_df.to_excel(writer, sheet_name="Shortlist", index=False)
        if backtest_df is not None and not backtest_df.empty:
            backtest_df.to_excel(writer, sheet_name="Backtest", index=False)
        if narratives:
            narr_df = pd.DataFrame(narratives)
            narr_df.to_excel(writer, sheet_name="Stock_narratives", index=False)
        guide_df = pd.DataFrame(_guide_sheet_data())
        guide_df.to_excel(writer, sheet_name="Guide", index=False, header=False)

    # Format workbook: column widths and text wrap for long content
    wb = openpyxl.load_workbook(REPORT_XLSX)
    for sheet_name in ["Sector_narrative", "Hypothesis", "Literature"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 90
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if "Stock_narratives" in wb.sheetnames:
        ws = wb["Stock_narratives"]
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            letter = get_column_letter(col_idx)
            if col and col[0].value and "narrative" in str(col[0].value).lower():
                ws.column_dimensions[letter].width = 70
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=col_idx)
                    c.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                ws.column_dimensions[letter].width = min(18, max(12, len(str(col[0].value or "")) + 1))
    if "Guide" in wb.sheetnames:
        ws = wb["Guide"]
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 75
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(REPORT_XLSX)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    sector_note = _read_text(SECTOR_NOTE_MD)
    sector_narrative = _read_text(SECTOR_NARRATIVE_MD)
    if not sector_narrative and "Definition and market size" in sector_note:
        sector_narrative = sector_note
    hypothesis = _read_text(HYPOTHESIS_MD)
    literature = _read_text(LITERATURE_MD)
    full_df = pd.read_csv(PHASE3_FULL_CSV) if PHASE3_FULL_CSV.exists() else pd.DataFrame()
    shortlist_df = pd.read_csv(PHASE3_SHORTLIST_CSV) if PHASE3_SHORTLIST_CSV.exists() else pd.DataFrame()
    backtest_df = pd.read_csv(PHASE4_BACKTEST_CSV) if PHASE4_BACKTEST_CSV.exists() else pd.DataFrame()
    narratives = _load_json(STOCK_NARRATIVES_JSON) if isinstance(_load_json(STOCK_NARRATIVES_JSON), list) else []
    as_of = full_df["AS_OF_DATE"].iloc[0] if full_df is not None and not full_df.empty and "AS_OF_DATE" in full_df.columns else date.today().isoformat()

    # MD
    md_content = build_md(
        sector_note, sector_narrative, hypothesis, literature,
        full_df, shortlist_df, backtest_df, narratives, as_of,
    )
    REPORT_MD.write_text(md_content, encoding="utf-8")
    print(f"  Wrote {REPORT_MD}")

    # HTML
    html_content = build_html(
        sector_narrative, hypothesis, literature,
        full_df, shortlist_df, backtest_df, narratives, as_of,
    )
    REPORT_HTML.write_text(html_content, encoding="utf-8")
    print(f"  Wrote {REPORT_HTML}")

    # XLSX
    build_xlsx(
        sector_narrative, hypothesis, literature,
        full_df, shortlist_df, backtest_df, narratives, as_of,
    )
    if REPORT_XLSX.exists():
        print(f"  Wrote {REPORT_XLSX}")

    print("Comprehensive report build complete.")


if __name__ == "__main__":
    main()
