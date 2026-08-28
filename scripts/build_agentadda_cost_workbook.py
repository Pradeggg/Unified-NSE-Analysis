#!/usr/bin/env python3
"""Build the AgentAdda.in product cost assessment workbook.

The source of truth is the CSV/memo pack in docs/superpowers/costing.
This script creates a polished .xlsx artifact for team review.
"""

from __future__ import annotations

import csv
from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
COST_DIR = ROOT / "docs" / "superpowers" / "costing"
OUT = COST_DIR / "AgentAdda_Product_Cost_Assessment_2026-08-23.xlsx"

ASSUMPTIONS = COST_DIR / "agentadda_cost_sheet_assumptions.csv"
LLM_COSTS = COST_DIR / "agentadda_llm_unit_costs.csv"
FEATURES = COST_DIR / "agentadda_feature_cogs_matrix.csv"
SCENARIOS = COST_DIR / "agentadda_subscription_scenarios.csv"
PAID_FEATURES = COST_DIR / "agentadda_paid_features_reports.csv"
REPORT_BUNDLES = COST_DIR / "agentadda_report_bundles_addons.csv"
DEPLOYMENT_COSTS = COST_DIR / "agentadda_deployment_architecture_costs.csv"
SUBSCRIBER_GROWTH = COST_DIR / "agentadda_subscriber_growth.csv"
BREAKEVEN_SUMMARY = COST_DIR / "agentadda_breakeven_summary.csv"


NAVY = "0F172A"
BLUE = "1D4ED8"
SKY = "DBEAFE"
GREEN = "DCFCE7"
AMBER = "FEF3C7"
RED = "FEE2E2"
GRAY = "F8FAFC"
MID = "E2E8F0"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="CBD5E1")
MED = Side(style="medium", color="64748B")


def rows_from_csv(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def set_title(ws, title: str, subtitle: str, last_col: int = 10) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, title)
    c.font = Font(bold=True, size=18, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    s = ws.cell(2, 1, subtitle)
    s.font = Font(italic=True, size=10, color="475569")
    s.fill = PatternFill("solid", fgColor=GRAY)
    s.alignment = Alignment(horizontal="center")


def style_range_as_table(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == min_row:
                cell.fill = PatternFill("solid", fgColor=BLUE)
                cell.font = Font(bold=True, color=WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GRAY)


def autosize(ws, min_width: int = 10, max_width: int = 42) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), max_width))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def currency_format(ws, cols: list[int], start_row: int = 1, end_row: int | None = None) -> None:
    end = end_row or ws.max_row
    for col in cols:
        for row in range(start_row, end + 1):
            ws.cell(row, col).number_format = '₹#,##0;[Red]-₹#,##0;"-"'


def percent_format(ws, cols: list[int], start_row: int = 1, end_row: int | None = None) -> None:
    end = end_row or ws.max_row
    for col in cols:
        for row in range(start_row, end + 1):
            ws.cell(row, col).number_format = '0.0%;[Red]-0.0%;"-"'


def write_csv_sheet(wb: Workbook, name: str, path: Path, title: str, subtitle: str) -> None:
    ws = wb.create_sheet(name)
    rows = rows_from_csv(path)
    set_title(ws, title, subtitle, last_col=max(8, len(rows[0])))
    start = 4
    for r, row in enumerate(rows, start):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    style_range_as_table(ws, start, start + len(rows) - 1, 1, len(rows[0]))
    ws.freeze_panes = f"A{start + 1}"
    autosize(ws)


def build_summary(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    set_title(
        ws,
        "AgentAdda.in Product Cost Assessment",
        "Subscription pricing, breakeven, and 20% profit assessment · prepared 2026-08-23",
        10,
    )
    ws.sheet_view.showGridLines = False

    cards = [
        ("Recommended default plan", "Plus ₹299/mo", "Best balance of value and margin"),
        ("100-user base revenue", "₹32,000/mo", "Blended Starter/Plus/Pro/Power mix"),
        ("Base-case operating profit", "₹7,040/mo", "After variable COGS and fixed beta costs"),
        ("Base-case margin", "22.0%", "Clears 20% target by ₹800 headroom"),
        ("Primary default LLM", "gpt-4o-mini", "Cost-efficient normal answer model"),
        ("Router / parser LLM", "gpt-5-nano", "Lowest-cost structured extraction layer"),
    ]
    positions = [(4, 1), (4, 4), (4, 7), (8, 1), (8, 4), (8, 7)]
    for (label, value, note), (row, col) in zip(cards, positions):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
        ws.cell(row, col, label).font = Font(bold=True, color="334155")
        ws.cell(row, col).fill = PatternFill("solid", fgColor=SKY)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 2)
        ws.cell(row + 1, col, value).font = Font(bold=True, size=16, color=NAVY)
        ws.cell(row + 1, col).fill = PatternFill("solid", fgColor=WHITE)
        ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 2)
        ws.cell(row + 2, col, note).font = Font(size=9, color="64748B")
        ws.cell(row + 2, col).fill = PatternFill("solid", fgColor=GRAY)
        for rr in range(row, row + 3):
            for cc in range(col, col + 3):
                ws.cell(rr, cc).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                ws.cell(rr, cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(13, 1, "Management Recommendation").font = Font(bold=True, size=13, color=WHITE)
    ws.cell(13, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=10)
    recommendations = [
        "Launch paid beta with Starter ₹149, Plus ₹299, Pro ₹799; keep Power ₹1,499 behind fair-use controls.",
        "Use gpt-5-nano for routing/extraction and gpt-4o-mini for default answers; reserve Luna/Terra for paid deep research.",
        "Do not use GPT-4, GPT-4 Turbo, or default GPT-4o for routine traffic; economics become unworkable.",
        "Before public launch: implement per-user token ledger, deep-research quotas, model-route logging, and budget circuit breakers.",
        "Validate production market-data licensing separately; real-time data can dominate LLM cost.",
    ]
    for i, text in enumerate(recommendations, 14):
        ws.cell(i, 1, f"{i-13}.")
        ws.cell(i, 2, text)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)
        ws.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(i, 1).font = Font(bold=True)

    ws.cell(21, 1, "Review tabs").font = Font(bold=True, color=WHITE)
    ws.cell(21, 1).fill = PatternFill("solid", fgColor=BLUE)
    ws.merge_cells(start_row=21, start_column=1, end_row=21, end_column=10)
    tabs = [
        ("Assumptions", "Editable pricing, usage, FX, and fixed-cost assumptions"),
        ("LLM Unit Costs", "Per-query and per-user monthly LLM economics"),
        ("Feature COGS", "Feature-by-feature cost drivers and controls"),
        ("Paid Features", "Paid feature/report packaging, tier gates, and usage controls"),
        ("Report Bundles", "Report packs and add-on monetization options"),
        ("Deployment Costs", "Cloudflare, FastAPI, Postgres, workers, email, monitoring"),
        ("Subscriber Growth", "Breakeven and profitability by subscriber count"),
        ("Subscription Model", "100-user plan mix, contribution, and margin"),
        ("Sensitivity", "All-Starter, all-Plus, all-Pro downside/upside cases"),
        ("Sources", "Pricing and product-scope references"),
    ]
    for i, (tab, desc) in enumerate(tabs, 22):
        ws.cell(i, 1, tab).font = Font(bold=True, color=NAVY)
        ws.cell(i, 2, desc)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 17
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    for row in range(1, 30):
        ws.row_dimensions[row].height = 22


def build_subscription_model(wb: Workbook) -> None:
    ws = wb.create_sheet("Subscription Model")
    set_title(ws, "100-User Subscription Model", "Base-case paid beta mix with breakeven and 20% profit target", 11)
    headers = [
        "Plan",
        "Users",
        "Price/User",
        "Revenue",
        "Variable COGS/User",
        "Variable COGS",
        "Contribution",
        "Fixed Cost Allocation",
        "Operating Profit",
        "Operating Margin",
        "Notes",
    ]
    start = 4
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    data = rows_from_csv(SCENARIOS)[1:6]
    for r, row in enumerate(data, start + 1):
        ws.cell(r, 1, row[0])
        ws.cell(r, 2, float(row[1]))
        ws.cell(r, 3, float(row[2]))
        ws.cell(r, 4, f"=B{r}*C{r}")
        ws.cell(r, 5, float(row[4]))
        ws.cell(r, 6, f"=B{r}*E{r}")
        ws.cell(r, 7, f"=D{r}-F{r}")
        ws.cell(r, 8, float(row[7]))
        ws.cell(r, 9, f"=G{r}-H{r}")
        ws.cell(r, 10, f"=I{r}/D{r}")
        ws.cell(r, 11, row[10])
    total_row = start + 1 + len(data)
    ws.cell(total_row, 1, "Formula check total")
    ws.cell(total_row, 2, f"=SUM(B{start+1}:B{total_row-1})")
    ws.cell(total_row, 3, f"=D{total_row}/B{total_row}")
    ws.cell(total_row, 4, f"=SUM(D{start+1}:D{total_row-1})")
    ws.cell(total_row, 5, f"=F{total_row}/B{total_row}")
    ws.cell(total_row, 6, f"=SUM(F{start+1}:F{total_row-1})")
    ws.cell(total_row, 7, f"=SUM(G{start+1}:G{total_row-1})")
    ws.cell(total_row, 8, f"=SUM(H{start+1}:H{total_row-1})")
    ws.cell(total_row, 9, f"=SUM(I{start+1}:I{total_row-1})")
    ws.cell(total_row, 10, f"=I{total_row}/D{total_row}")
    ws.cell(total_row, 11, "All formulas driven from rows above")

    target_row = total_row + 3
    ws.cell(target_row, 1, "Breakeven / 20% Margin Check").font = Font(bold=True, color=WHITE)
    ws.cell(target_row, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=6)
    checks = [
        ("Total cost", f"=F{total_row}+H{total_row}"),
        ("Breakeven revenue", f"=B{target_row+1}"),
        ("Revenue required for 20% operating margin", f"=B{target_row+1}/0.8"),
        ("Actual revenue", f"=D{total_row}"),
        ("Headroom vs 20% target", f"=B{target_row+4}-B{target_row+3}"),
    ]
    for i, (label, formula) in enumerate(checks, target_row + 1):
        ws.cell(i, 1, label)
        ws.cell(i, 2, formula)
        ws.cell(i, 2).number_format = '₹#,##0;[Red]-₹#,##0;"-"'

    style_range_as_table(ws, start, total_row, 1, len(headers))
    for cell in ws[total_row]:
        cell.fill = PatternFill("solid", fgColor=AMBER)
        cell.font = Font(bold=True, color=NAVY)
        cell.border = Border(top=MED, bottom=MED, left=THIN, right=THIN)

    currency_format(ws, [3, 4, 5, 6, 7, 8, 9], start + 1, total_row)
    percent_format(ws, [10], start + 1, total_row)
    ws.freeze_panes = "A5"
    autosize(ws, max_width=36)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Revenue vs Variable COGS by Plan"
    chart.y_axis.title = "INR / month"
    chart.x_axis.title = "Plan"
    data_ref = Reference(ws, min_col=4, max_col=6, min_row=start, max_row=total_row - 1)
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=total_row - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 16
    ws.add_chart(chart, "A19")


def build_sensitivity(wb: Workbook) -> None:
    ws = wb.create_sheet("Sensitivity")
    set_title(ws, "Subscription Sensitivity", "Downside/upside cases for 100 users", 10)
    headers = ["Scenario", "Users", "Price/User", "Revenue", "Variable COGS/User", "Variable COGS", "Fixed Cost", "Profit", "Margin", "Conclusion"]
    start = 4
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    data = rows_from_csv(SCENARIOS)[6:]
    for r, row in enumerate(data, start + 1):
        ws.cell(r, 1, row[0])
        ws.cell(r, 2, float(row[1]))
        ws.cell(r, 3, float(row[2]))
        ws.cell(r, 4, f"=B{r}*C{r}")
        ws.cell(r, 5, float(row[4]))
        ws.cell(r, 6, f"=B{r}*E{r}")
        ws.cell(r, 7, float(row[7]))
        ws.cell(r, 8, f"=D{r}-F{r}-G{r}")
        ws.cell(r, 9, f"=H{r}/D{r}")
        ws.cell(r, 10, row[10])
    style_range_as_table(ws, start, start + len(data), 1, len(headers))
    currency_format(ws, [3, 4, 5, 6, 7, 8], start + 1, start + len(data))
    percent_format(ws, [9], start + 1, start + len(data))
    autosize(ws, max_width=44)

    chart = BarChart()
    chart.title = "Operating Profit by Scenario"
    chart.y_axis.title = "INR / month"
    chart.x_axis.title = "Scenario"
    chart.add_data(Reference(ws, min_col=8, min_row=start, max_row=start + len(data)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(data)))
    chart.height = 7
    chart.width = 15
    ws.add_chart(chart, "A11")


def build_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources")
    set_title(ws, "Sources and Caveats", "Provider pricing and internal product-scope references", 4)
    rows = [
        ["Source", "URL / Path", "Use", "Notes"],
        ["OpenAI model catalog", "https://developers.openai.com/api/docs/models", "OpenAI model selection/pricing", "Fetched 2026-08-23"],
        ["OpenAI gpt-4o-mini", "https://developers.openai.com/api/docs/models/gpt-4o-mini", "Older low-cost model pricing", "Used in default paid route"],
        ["OpenAI gpt-4o", "https://developers.openai.com/api/docs/models/gpt-4o", "Premium older model pricing", "Avoid routine use"],
        ["OpenAI gpt-4", "https://developers.openai.com/api/docs/models/gpt-4", "Legacy GPT-4 pricing", "Not subscription viable"],
        ["OpenAI deprecations", "https://developers.openai.com/api/docs/deprecations", "Legacy model lifecycle risk", "Check before hardcoding snapshots"],
        ["DeepSeek pricing", "https://api-docs.deepseek.com/quick_start/pricing/", "Batch fallback pricing", "Off-peak price used"],
        ["Sarvam pricing", "https://docs.sarvam.ai/api/getting-started/pricing", "Indic language and speech pricing", "Presentation layer candidate"],
        ["xAI pricing", "https://docs.x.ai/developers/pricing", "Grok optional pricing", "Not default-cost viable"],
        ["Wise FX", "https://wise.com/us/currency-converter/usd-to-inr-rate/history", "USD/INR conversion", "₹95.66/USD"],
        ["Talk 2 Stocks spec", "docs/superpowers/specs/2026-08-23-talk-2-stocks-comprehensive-product-design.md", "Internal product scope", "Local repo file"],
        ["Agent Adda terminal", "nse_agent.py", "Feature registry", "Local repo file"],
    ]
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    style_range_as_table(ws, 4, 4 + len(rows) - 1, 1, 4)
    autosize(ws, max_width=55)


def finalize(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.font:
                    font = copy(cell.font)
                    font.name = "Aptos"
                    cell.font = font
                else:
                    cell.font = Font(name="Aptos")
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0


def main() -> None:
    wb = Workbook()
    build_summary(wb)
    write_csv_sheet(wb, "Assumptions", ASSUMPTIONS, "Assumptions", "Editable source assumptions used by the model")
    write_csv_sheet(wb, "LLM Unit Costs", LLM_COSTS, "LLM Unit Costs", "Per-query and per-user LLM economics")
    currency_format(wb["LLM Unit Costs"], [5, 6, 7, 8, 9], 5)
    write_csv_sheet(wb, "Feature COGS", FEATURES, "Feature COGS Matrix", "Feature-level cost drivers, routes, and controls")
    write_csv_sheet(wb, "Paid Features", PAID_FEATURES, "Paid Features and Reports", "Monetizable product surfaces, report packs, and tier gates")
    write_csv_sheet(wb, "Report Bundles", REPORT_BUNDLES, "Report Bundles and Add-ons", "Paid report packs and optional monetization extensions")
    write_csv_sheet(wb, "Deployment Costs", DEPLOYMENT_COSTS, "Deployment Architecture Costs", "Current Cloudflare + FastAPI + PostgreSQL + worker cost mapping")
    write_csv_sheet(wb, "Breakeven Summary", BREAKEVEN_SUMMARY, "Breakeven Summary", "Subscribers required for breakeven and 20% operating margin")
    write_csv_sheet(wb, "Subscriber Growth", SUBSCRIBER_GROWTH, "Subscriber Growth Model", "Profitability curve across lean beta and production infrastructure cases")
    build_subscription_model(wb)
    build_sensitivity(wb)
    build_sources(wb)
    finalize(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
